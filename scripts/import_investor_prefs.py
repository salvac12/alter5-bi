#!/usr/bin/env python3
"""
===============================================================
  Alter5 BI -- Import Investor Preferences from Excel
===============================================================

  Reads preferencias_inversores_v1.xlsx and merges investor
  preference fields into companies_full.json enrichment data.

  Fields imported:
    sentiment, inv_phase, ticket_size, asset_types,
    inv_criteria, next_action, deals_mentioned, inv_subtipo

  Fields NOT imported (already exist from email pipeline):
    Contactos Clave, Buzones, N Emails Analizados

  Usage:
    python scripts/import_investor_prefs.py                         # import all
    python scripts/import_investor_prefs.py --dry-run               # preview
    python scripts/import_investor_prefs.py --stats                 # stats only
    python scripts/import_investor_prefs.py --domain bankinter.com  # single
===============================================================
"""

import argparse
import json
import os
import sys
from datetime import datetime

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, SCRIPT_DIR)

from import_mailbox import get_data_paths

# Load .env file if present
def _load_dotenv():
    d = PROJECT_DIR
    for _ in range(6):
        candidate = os.path.join(d, ".env")
        if os.path.exists(candidate):
            with open(candidate) as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, val = line.partition("=")
                    key, val = key.strip(), val.strip().strip('"').strip("'")
                    if key and key not in os.environ:
                        os.environ[key] = val
            break
        d = os.path.dirname(d)

_load_dotenv()

# Default Excel path
DEFAULT_EXCEL = os.path.join(os.path.expanduser("~"), "Downloads", "preferencias_inversores_v1.xlsx")

# Column indices (0-based) in "Todas las Preferencias" sheet
COL_EMPRESA = 0
COL_DOMINIO = 1
COL_SUBTIPO = 2
COL_SENTIMIENTO = 3
COL_FASE = 4
COL_TICKET = 5
COL_GEOGRAFIAS = 6
COL_TECNOLOGIAS = 7
COL_TIPO_ACTIVO = 8
COL_CRITERIOS = 9
COL_DEALS = 10
# COL_CONTACTOS = 11  # skip
COL_SIGUIENTE = 12
# COL_NEMAILS = 13    # skip
# COL_BUZONES = 14    # skip


def parse_multi(val):
    """Parse pipe-separated multi-value field into list."""
    if not val:
        return []
    return [v.strip() for v in str(val).split("|") if v.strip()]


def read_excel(path):
    """Read the Excel file and return list of dicts."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Todas las Preferencias"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        print(f"  Warning: sheet 'Todas las Preferencias' not found, using '{sheet_name}'")

    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        domain = row[COL_DOMINIO]
        if not domain:
            continue
        domain = str(domain).strip().lower()
        if not domain or domain == "dominio":
            continue

        sentiment = str(row[COL_SENTIMIENTO] or "").strip() or None
        inv_phase = str(row[COL_FASE] or "").strip() or None
        ticket_raw = str(row[COL_TICKET] or "").strip()
        ticket_size = ticket_raw if ticket_raw and ticket_raw != "desconocido" else None

        asset_types = parse_multi(row[COL_TIPO_ACTIVO])
        deals = parse_multi(row[COL_DEALS])
        inv_criteria = str(row[COL_CRITERIOS] or "").strip() or None
        next_action = str(row[COL_SIGUIENTE] or "").strip() or None
        inv_subtipo = str(row[COL_SUBTIPO] or "").strip() or None

        rows.append({
            "domain": domain,
            "empresa": str(row[COL_EMPRESA] or "").strip(),
            "sentiment": sentiment,
            "inv_phase": inv_phase,
            "ticket_size": ticket_size,
            "asset_types": asset_types,
            "inv_criteria": inv_criteria,
            "next_action": next_action,
            "deals_mentioned": deals,
            "inv_subtipo": inv_subtipo,
        })

    wb.close()
    return rows


def main():
    parser = argparse.ArgumentParser(description="Import investor preferences from Excel")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Path to Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    parser.add_argument("--stats", action="store_true", help="Show stats only")
    parser.add_argument("--domain", help="Process single domain")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"ERROR: Excel file not found: {args.excel}")
        sys.exit(1)

    print(f"Reading Excel: {args.excel}")
    excel_rows = read_excel(args.excel)
    print(f"  {len(excel_rows)} rows read from Excel")

    # Load companies_full.json
    paths = get_data_paths(PROJECT_DIR)
    full_path = paths["full"]
    print(f"Loading: {full_path}")
    with open(full_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    companies = data.get("companies", {})
    print(f"  {len(companies)} companies in database")

    # Build domain lookup
    excel_by_domain = {}
    for row in excel_rows:
        excel_by_domain[row["domain"]] = row

    # Stats
    matched = 0
    unmatched_domains = []
    updated = 0
    sentiments = {}
    phases = {}
    has_ticket = 0
    has_criteria = 0
    has_next = 0
    has_deals = 0

    now = datetime.now(tz=__import__('datetime').timezone.utc).isoformat().replace("+00:00", "Z")

    for row in excel_rows:
        domain = row["domain"]

        # Track stats
        if row["sentiment"]:
            sentiments[row["sentiment"]] = sentiments.get(row["sentiment"], 0) + 1
        if row["inv_phase"]:
            phases[row["inv_phase"]] = phases.get(row["inv_phase"], 0) + 1
        if row["ticket_size"]:
            has_ticket += 1
        if row["inv_criteria"]:
            has_criteria += 1
        if row["next_action"]:
            has_next += 1
        if row["deals_mentioned"]:
            has_deals += 1

        if args.domain and domain != args.domain.lower():
            continue

        if domain not in companies:
            unmatched_domains.append(domain)
            continue

        matched += 1
        comp = companies[domain]
        enrichment = comp.setdefault("enrichment", {})

        # Write investor preference fields
        fields_to_write = {
            "sentiment": row["sentiment"],
            "inv_phase": row["inv_phase"],
            "ticket_size": row["ticket_size"],
            "inv_subtipo": row["inv_subtipo"],
            "inv_criteria": row["inv_criteria"],
            "next_action": row["next_action"],
        }

        changed = False
        for key, val in fields_to_write.items():
            if val is not None:
                if enrichment.get(key) != val:
                    enrichment[key] = val
                    changed = True

        # Array fields
        if row["asset_types"]:
            if enrichment.get("asset_types") != row["asset_types"]:
                enrichment["asset_types"] = row["asset_types"]
                changed = True
        if row["deals_mentioned"]:
            if enrichment.get("deals_mentioned") != row["deals_mentioned"]:
                enrichment["deals_mentioned"] = row["deals_mentioned"]
                changed = True

        if changed:
            enrichment["_inv_source"] = "excel_v1"
            enrichment["_inv_updated_at"] = now
            updated += 1

    # Print stats
    print(f"\n{'='*50}")
    print(f"  Excel rows:     {len(excel_rows)}")
    print(f"  Matched:        {matched}")
    print(f"  Unmatched:      {len(unmatched_domains)}")
    print(f"  Updated:        {updated}")
    print(f"\n  Sentimientos:")
    for k, v in sorted(sentiments.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}")
    print(f"\n  Fases:")
    for k, v in sorted(phases.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}")
    print(f"\n  Con ticket size: {has_ticket}")
    print(f"  Con criterios:   {has_criteria}")
    print(f"  Con next action: {has_next}")
    print(f"  Con deals:       {has_deals}")

    if unmatched_domains and not args.stats:
        print(f"\n  Top unmatched domains:")
        for d in unmatched_domains[:20]:
            print(f"    - {d}")
        if len(unmatched_domains) > 20:
            print(f"    ... and {len(unmatched_domains) - 20} more")

    if args.stats:
        return

    if args.dry_run:
        print(f"\n  DRY RUN — no files written")
        return

    if updated == 0:
        print(f"\n  No changes to write")
        return

    # Write back
    print(f"\nWriting {full_path}...")
    with open(full_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  Done! {updated} companies updated.")

    # Also regenerate compact format
    compact_path = paths["compact"]
    print(f"Regenerating compact: {compact_path}...")
    from import_mailbox import export_to_compact
    all_companies = data.get("companies", {})
    compact = export_to_compact(all_companies)
    with open(compact_path, "w", encoding="utf-8") as f:
        json.dump(compact, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  Done!")


if __name__ == "__main__":
    main()
